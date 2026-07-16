"""HTTP-Integrationstest fuer Exportpakete mit temporaerer SQLite-DB."""
import io, json, os, pathlib, socket, subprocess, sys, tempfile, time, unittest
import urllib.error, urllib.request
from openpyxl import load_workbook

TEMP_DB = tempfile.TemporaryDirectory(prefix="finanz-export-")
os.environ["FINANZ_DB"] = str(pathlib.Path(TEMP_DB.name) / "test.db")
ROOT = pathlib.Path(__file__).resolve().parents[1]

class ExportTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        with socket.socket() as sock:
            sock.bind(("127.0.0.1", 0)); cls.port = sock.getsockname()[1]
        cls.base = f"http://127.0.0.1:{cls.port}"
        cls.server = subprocess.Popen(
            [sys.executable, "-m", "uvicorn", "app.main:app", "--host", "127.0.0.1",
             "--port", str(cls.port), "--log-level", "warning"], cwd=ROOT,
            env=os.environ.copy(), stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
        deadline = time.monotonic() + 90
        while time.monotonic() < deadline:
            try:
                cls.raw("GET", "/api/health"); break
            except Exception:
                if cls.server.poll() is not None: raise RuntimeError("uvicorn beendet")
                time.sleep(.2)
        else: raise RuntimeError("uvicorn nicht bereit")
        cls.sparten = cls.req("GET", "/api/sparten")[:2]
        cls.kategorien = [cls.req("POST", "/api/kategorien", 201, {
            "sparte_id": s["id"], "name": f"Kategorie <{i}>", "richtung": "beides"})
            for i, s in enumerate(cls.sparten, 1)]
        daten = (
            (0, "2026-01-15", "einnahme", 123456, "Jahreseinnahme"),
            (0, "2026-02-15", "ausgabe", 2345, "Material"),
            (1, "2026-01-20", "ausgabe", 5000, "Andere Sparte"),
            (0, "2025-12-31", "einnahme", 9999, "Ausserhalb"))
        for index, datum, typ, betrag, text in daten:
            cls.req("POST", "/api/buchungen", 201, {
                "sparte_id": cls.sparten[index]["id"], "datum": datum, "typ": typ, "text": text,
                "zeilen": [{"kategorie_id": cls.kategorien[index]["id"],
                            "betrag_cent": betrag}]})

    @classmethod
    def tearDownClass(cls):
        if cls.server.poll() is None:
            cls.server.terminate(); cls.server.wait(timeout=10)
        TEMP_DB.cleanup()

    @classmethod
    def raw(cls, method, path, body=None):
        data = json.dumps(body).encode() if body is not None else None
        req = urllib.request.Request(cls.base + path, data=data, method=method,
            headers={"Content-Type": "application/json"} if data else {})
        try:
            with urllib.request.urlopen(req, timeout=15) as res:
                return res.status, res.headers, res.read()
        except urllib.error.HTTPError as error:
            return error.code, error.headers, error.read()

    @classmethod
    def req(cls, method, path, expected=200, body=None):
        status, _headers, content = cls.raw(method, path, body)
        if status != expected:
            raise AssertionError(f"{path}: {status} != {expected}: {content!r}")
        return json.loads(content) if content else None

    def test_xlsx_hat_drei_blaetter_numerische_euro_und_filter(self):
        sid = self.sparten[0]["id"]
        status, headers, content = self.raw(
            "GET", f"/api/export/xlsx?von=2026-01-01&bis=2026-12-31&sparte_id={sid}")
        self.assertEqual(status, 200, content.decode(errors="replace"))
        self.assertIn("spreadsheetml", headers.get_content_type())
        self.assertIn("attachment", headers.get("Content-Disposition", ""))
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        self.assertEqual(workbook.sheetnames, ["Buchungen", "Monatssummen", "Kategorien"])
        buchungen = workbook["Buchungen"]
        self.assertEqual([cell.value for cell in buchungen[1]],
            ["Datum", "Sparte", "Kategorie", "Typ", "Text", "Kontakt", "Betrag \u20ac"])
        self.assertEqual(buchungen.max_row, 3)
        self.assertEqual({buchungen.cell(row, 2).value for row in range(2, 4)},
                         {self.sparten[0]["name"]})
        for row in range(2, 4):
            self.assertIsInstance(buchungen.cell(row, 7).value, (int, float))
            self.assertIn("0,00", buchungen.cell(row, 7).number_format)
        self.assertEqual(buchungen.cell(2, 7).value, 1234.56)
        monate = workbook["Monatssummen"]
        self.assertEqual(monate.cell(2, 2).value, 1234.56)
        self.assertEqual(monate.cell(2, 3).value, 0)
        self.assertEqual(monate.cell(3, 3).value, 23.45)
        self.assertTrue(all(isinstance(monate.cell(r, c).value, (int, float))
            for r in range(2, monate.max_row + 1) for c in range(2, 5)))
        kategorien = workbook["Kategorien"]
        self.assertIn("Gesamt", [kategorien.cell(r, 2).value
                                  for r in range(2, kategorien.max_row + 1)])
        self.assertTrue(all(isinstance(kategorien.cell(r, c).value, (int, float))
            for r in range(2, kategorien.max_row + 1) for c in range(3, 6)))

    def test_xlsx_neutralisiert_formelpraefixe_in_textzellen(self):
        sid = self.sparten[0]["id"]
        dangerous = '=HYPERLINK("https://example.invalid","Klick")'
        self.req("POST", "/api/buchungen", 201, {
            "sparte_id": sid,
            "datum": "2027-01-01",
            "typ": "ausgabe",
            "text": dangerous,
            "zeilen": [{
                "kategorie_id": self.kategorien[0]["id"],
                "betrag_cent": 100,
            }],
        })

        status, _headers, content = self.raw(
            "GET", f"/api/export/xlsx?von=2027-01-01&bis=2027-12-31&sparte_id={sid}")
        self.assertEqual(status, 200, content.decode(errors="replace"))
        cell = load_workbook(io.BytesIO(content), data_only=False)["Buchungen"]["E2"]
        self.assertEqual(cell.data_type, "s")
        self.assertEqual(cell.value, "'" + dangerous)

    def test_bericht_ist_druckbar_gefiltert_formatiert_und_escaped(self):
        sid = self.sparten[0]["id"]
        status, headers, content = self.raw("GET", f"/export/bericht?jahr=2026&sparte_id={sid}")
        html = content.decode("utf-8")
        self.assertEqual(status, 200, html)
        self.assertEqual(headers.get_content_type(), "text/html")
        for marker in ("Jahresbericht 2026", "1.234,56 \u20ac", "23,45 \u20ac", "1.211,11 \u20ac",
                       "window.print()", "@media print", "Drucken / Als PDF speichern",
                       self.sparten[0]["name"], "Kategorie &lt;1&gt;"):
            self.assertIn(marker, html)
        self.assertNotIn(self.sparten[1]["name"], html)
        self.assertNotIn("Kategorie <1>", html)

    def test_fehlerhafte_parameter_werden_deutsch_abgewiesen(self):
        for path in ("/api/export/xlsx?von=2026-13-01",
                     "/api/export/xlsx?von=2026-12-31&bis=2026-01-01",
                     "/export/bericht?jahr=20xx"):
            status, _headers, content = self.raw("GET", path)
            self.assertEqual(status, 400, (path, content))
            self.assertIn("detail", json.loads(content))
        status, _headers, content = self.raw("GET", "/export/bericht?jahr=2026&sparte_id=999999")
        self.assertEqual(status, 404, content)
        self.assertIn("Sparte", json.loads(content)["detail"])

    def test_studio_export_vertrag(self):
        html = (ROOT / "static-studio" / "index.html").read_text(encoding="utf-8")
        js = (ROOT / "static-studio" / "app.js").read_text(encoding="utf-8")
        for marker in ("ex-von", "ex-bis", "ex-xlsx", "ex-bericht"):
            self.assertIn(f'id="{marker}"', html)
        self.assertIn("/export/xlsx", js)
        self.assertIn("/export/bericht", js)

if __name__ == "__main__": unittest.main(verbosity=2)
