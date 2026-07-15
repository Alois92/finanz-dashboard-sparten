"""Excel-Kassabuch-Import (Altdaten-Migration).

Liest die Kassabuch-Excel-Dateien des Benutzers (Muster: reference/
"Freizeit 1.0 Original.xls"): 12 Monatsblaetter + "Jahreskassabuch".
Aufbau je Monatsblatt:

  Zeile 6 (Index 5):  Kopf - Spalte A 'Dat.', C Text, D EINNAHMEN,
                      E AUSGABEN, ab Spalte L (Index 11) die Kategorienamen
  Zeile 8 (Index 7):  erste Datenzeile
  Fusszeilen:         'BRUTTO-EINNAHMEN' usw. in Spalte C -> Ende der Daten

Das Blatt "Jahreskassabuch" enthaelt nur Summen und wird uebersprungen.
Unterstuetzt .xls (xlrd) und .xlsx (openpyxl).

Endpunkte:
  POST /api/import/excel  - multipart: datei, sparte_id, modus=pruefen|einspielen
                            'pruefen' aendert NICHTS und liefert nur den Bericht.
"""
import datetime as dt
import io
import sqlite3
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile

from ..db import db_dep

router = APIRouter(tags=["import"])

KATEGORIE_SPALTE_AB = 11          # Index der ersten Kategorie-Spalte ("L")
FALLBACK_KATEGORIE = "Nicht zugeordnet"
MAX_WARNUNGEN = 50


# ---------------------------------------------------------------------------
# Datei lesen: .xls / .xlsx auf ein gemeinsames Zellenraster normalisieren
# ---------------------------------------------------------------------------

def _lade_blaetter(rohdaten: bytes, dateiname: str) -> list[tuple[str, list[list]]]:
    """Liefert [(blattname, zeilen)] - Zellen als str/float/datetime/None."""
    name = (dateiname or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(rohdaten), data_only=True,
                                    read_only=True)
        blaetter = []
        for ws in wb.worksheets:
            zeilen = [[c for c in row] for row in ws.iter_rows(values_only=True)]
            blaetter.append((ws.title, zeilen))
        wb.close()
        return blaetter
    # .xls (BIFF) ueber xlrd; Datumszellen als datetime aufloesen
    import xlrd
    try:
        wb = xlrd.open_workbook(file_contents=rohdaten)
    except xlrd.XLRDError as e:
        raise HTTPException(400, f"Datei kann nicht gelesen werden: {e}")
    blaetter = []
    for sh in wb.sheets():
        zeilen = []
        for r in range(sh.nrows):
            zeile = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        zeile.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
                    except Exception:
                        zeile.append(None)
                elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    zeile.append(None)
                else:
                    zeile.append(cell.value)
            zeilen.append(zeile)
        blaetter.append((sh.name, zeilen))
    return blaetter


# ---------------------------------------------------------------------------
# Zellwert-Helfer
# ---------------------------------------------------------------------------

def _text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _betrag_cent(v) -> Optional[int]:
    """Zahl -> Cent; leere/Nicht-Zahlen -> None; 0 -> None (leere Formelzelle)."""
    if v is None or isinstance(v, (dt.datetime, dt.date)):
        return None
    if isinstance(v, (int, float)):
        cent = round(float(v) * 100)
        return cent if cent != 0 else None
    s = str(v).strip().replace("€", "").replace(" ", "")
    if not s:
        return None
    try:
        cent = round(float(s.replace(".", "").replace(",", ".")) * 100)
        return cent if cent != 0 else None
    except ValueError:
        return None


def _datum(v, datei_jahr: Optional[int]) -> Optional[str]:
    """Zelle -> ISO-Datum. Excel liefert datetime; Strings 'T.M.(JJJJ)' werden
    geparst (fehlt das Jahr, wird datei_jahr ergaenzt)."""
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, (int, float)) and v > 0:
        # xls-Serial, das xlrd nicht als Datum erkannt hat (Zellformat fehlt)
        try:
            basis = dt.date(1899, 12, 30)
            return (basis + dt.timedelta(days=int(v))).isoformat()
        except OverflowError:
            return None
    s = _text(v)
    if not s:
        return None
    teile = s.rstrip(".").split(".")
    try:
        if len(teile) == 3 and teile[2]:
            jahr = int(teile[2])
            if jahr < 100:
                jahr += 2000
            return dt.date(jahr, int(teile[1]), int(teile[0])).isoformat()
        if len(teile) >= 2 and datei_jahr:
            return dt.date(datei_jahr, int(teile[1]), int(teile[0])).isoformat()
    except (ValueError, IndexError):
        return None
    return None


# ---------------------------------------------------------------------------
# Kassabuch-Parser
# ---------------------------------------------------------------------------

def _parse_kassabuch(blaetter: list[tuple[str, list[list]]]) -> dict:
    """Extrahiert Buchungs-Kandidaten aus allen Monatsblaettern.

    Rueckgabe: {"posten": [...], "warnungen": [...], "monate_mit_daten": n}
    Ein Posten: {datum, text, beleg_nr, typ, betrag_cent,
                 zeilen: [(kategorie_name, betrag_cent)], blatt, zeile}
    """
    posten: list[dict] = []
    warnungen: list[str] = []
    monate_mit_daten = 0

    def warne(msg: str):
        if len(warnungen) < MAX_WARNUNGEN:
            warnungen.append(msg)

    for blattname, zeilen in blaetter:
        if "jahres" in blattname.lower():
            continue  # Summenblatt, sonst zaehlt alles doppelt

        # Kopfzeile finden ('Dat.' in Spalte A)
        kopf_idx = None
        for i, z in enumerate(zeilen[:15]):
            if z and _text(z[0] if len(z) > 0 else "").lower().startswith("dat"):
                kopf_idx = i
                break
        if kopf_idx is None:
            continue  # kein Kassabuch-Blatt

        kopf = zeilen[kopf_idx]
        kategorien: list[tuple[int, str]] = []
        for c in range(KATEGORIE_SPALTE_AB, len(kopf)):
            name = _text(kopf[c])
            if name:
                kategorien.append((c, name))

        # Jahr aus Kopfbereich (Zelle 'Jahr:' daneben) - nur als Fallback
        datei_jahr = None
        for z in zeilen[:5]:
            for c, v in enumerate(z or []):
                if _text(v).lower().startswith("jahr") and c + 1 < len(z):
                    j = _betrag_cent(z[c + 1])
                    if j and 190000 <= j // 100 <= 210000:
                        datei_jahr = j // 100

        blatt_hat_daten = False
        for r in range(kopf_idx + 2, len(zeilen)):
            z = zeilen[r]
            if not z:
                continue
            marker = _text(z[2] if len(z) > 2 else "").upper()
            if marker.startswith(("BRUTTO", "MWST", "VORST", "GUTSCHRIFT")):
                break  # Fusszeilen erreicht

            einnahme = _betrag_cent(z[3] if len(z) > 3 else None)
            ausgabe = _betrag_cent(z[4] if len(z) > 4 else None)
            if einnahme is None and ausgabe is None:
                continue  # leere Vorlagenzeile

            datum = _datum(z[0] if len(z) > 0 else None, datei_jahr)
            text = _text(z[2] if len(z) > 2 else "")
            beleg_nr = _text(z[1] if len(z) > 1 else "")
            if datum is None:
                warne(f"{blattname} Zeile {r + 1}: Betrag ohne Datum - uebersprungen")
                continue

            # Kategorie-Spalten dieser Zeile einsammeln
            kat_werte: list[tuple[str, int]] = []
            for c, name in kategorien:
                w = _betrag_cent(z[c] if len(z) > c else None)
                if w is not None:
                    kat_werte.append((name, abs(w)))

            for typ, betrag in (("einnahme", einnahme), ("ausgabe", ausgabe)):
                if betrag is None:
                    continue
                betrag = abs(betrag)
                # Zeilen: passende Kategorie-Werte; sonst Fallback in voller Hoehe
                if len(kat_werte) == 1:
                    kzeilen = [(kat_werte[0][0], betrag)]
                elif len(kat_werte) > 1:
                    summe = sum(w for _, w in kat_werte)
                    if summe == betrag:
                        kzeilen = list(kat_werte)
                    else:
                        # Split passt nicht zum Betrag (z. B. Einnahme+Ausgabe in
                        # einer Zeile): exakte Einzelspalte suchen, sonst Fallback.
                        exakt = [kv for kv in kat_werte if kv[1] == betrag]
                        if len(exakt) == 1:
                            kzeilen = [exakt[0]]
                        else:
                            kzeilen = [(FALLBACK_KATEGORIE, betrag)]
                            warne(f"{blattname} Zeile {r + 1}: Kontierung "
                                  f"({summe / 100:.2f}) passt nicht zum Betrag "
                                  f"({betrag / 100:.2f}) - '{FALLBACK_KATEGORIE}'")
                else:
                    kzeilen = [(FALLBACK_KATEGORIE, betrag)]
                    warne(f"{blattname} Zeile {r + 1}: keine Kontierungsspalte "
                          f"befuellt - '{FALLBACK_KATEGORIE}'")

                posten.append({
                    "datum": datum, "text": text, "beleg_nr": beleg_nr,
                    "typ": typ, "betrag_cent": betrag, "zeilen": kzeilen,
                    "blatt": blattname, "zeile": r + 1,
                })
                blatt_hat_daten = True

        if blatt_hat_daten:
            monate_mit_daten += 1

    return {"posten": posten, "warnungen": warnungen,
            "monate_mit_daten": monate_mit_daten}


# ---------------------------------------------------------------------------
# Endpoint
# ---------------------------------------------------------------------------

@router.post("/import/excel")
def import_excel(
    sparte_id: int = Form(...),
    modus: str = Form("pruefen"),
    datei: UploadFile = File(...),
    con: sqlite3.Connection = Depends(db_dep),
):
    if modus not in ("pruefen", "einspielen"):
        raise HTTPException(400, "modus muss 'pruefen' oder 'einspielen' sein")
    if not con.execute("SELECT 1 FROM sparte WHERE id = ?", (sparte_id,)).fetchone():
        raise HTTPException(404, "Sparte nicht gefunden")

    rohdaten = datei.file.read()
    if not rohdaten:
        raise HTTPException(400, "Datei ist leer")

    ergebnis = _parse_kassabuch(_lade_blaetter(rohdaten, datei.filename or ""))
    posten = ergebnis["posten"]
    warnungen = ergebnis["warnungen"]

    # Vorhandene Kategorien der Sparte (Name -> id, case-insensitiv)
    kat_vorhanden = {
        r["name"].strip().lower(): r["id"]
        for r in con.execute(
            "SELECT id, name FROM kategorie WHERE sparte_id = ? AND aktiv = 1",
            (sparte_id,)).fetchall()
    }
    benoetigt: dict[str, str] = {}   # lower -> Originalname
    for p in posten:
        for name, _ in p["zeilen"]:
            benoetigt.setdefault(name.strip().lower(), name.strip())
    neue_kategorien = sorted(orig for low, orig in benoetigt.items()
                             if low not in kat_vorhanden)

    # Dubletten: identische Buchung existiert schon (macht Wiederholungen harmlos)
    def ist_dublette(p: dict) -> bool:
        return con.execute(
            "SELECT 1 FROM buchung WHERE sparte_id = ? AND datum = ? AND typ = ? "
            "AND betrag_cent = ? AND IFNULL(text,'') = ? LIMIT 1",
            (sparte_id, p["datum"], p["typ"], p["betrag_cent"], p["text"] or ""),
        ).fetchone() is not None

    duplikate = [p for p in posten if ist_dublette(p)]
    neu = [p for p in posten if p not in duplikate]

    # Zusammenfassung je Kategorie
    je_kategorie: dict[str, dict] = {}
    for p in neu:
        for name, betrag in p["zeilen"]:
            eintrag = je_kategorie.setdefault(
                name, {"kategorie": name, "anzahl": 0, "einnahmen_cent": 0,
                       "ausgaben_cent": 0,
                       "neu": name.strip().lower() not in kat_vorhanden})
            eintrag["anzahl"] += 1
            eintrag["einnahmen_cent" if p["typ"] == "einnahme"
                    else "ausgaben_cent"] += betrag

    bericht = {
        "dateiname": datei.filename,
        "modus": modus,
        "sparte_id": sparte_id,
        "monate_mit_daten": ergebnis["monate_mit_daten"],
        "buchungen_neu": len(neu),
        "duplikate": len(duplikate),
        "einnahmen_cent": sum(p["betrag_cent"] for p in neu if p["typ"] == "einnahme"),
        "ausgaben_cent": sum(p["betrag_cent"] for p in neu if p["typ"] == "ausgabe"),
        "zeitraum": ({"von": min(p["datum"] for p in neu),
                      "bis": max(p["datum"] for p in neu)} if neu else None),
        "neue_kategorien": neue_kategorien,
        "je_kategorie": sorted(je_kategorie.values(),
                               key=lambda k: -(k["einnahmen_cent"] + k["ausgaben_cent"])),
        "warnungen": warnungen,
        "eingespielt": 0,
    }
    if modus == "pruefen":
        return bericht

    # ---- einspielen ----
    try:
        for low, orig in sorted(benoetigt.items()):
            if low not in kat_vorhanden:
                cur = con.execute(
                    "INSERT INTO kategorie(sparte_id, name, richtung) "
                    "VALUES(?,?, 'beides')", (sparte_id, orig))
                kat_vorhanden[low] = cur.lastrowid
        for p in neu:
            cur = con.execute(
                "INSERT INTO buchung(sparte_id, datum, typ, zahlungsart, "
                "buchungsstatus, text, notiz) VALUES(?,?,?,'bar','zugeordnet',?,?)",
                (sparte_id, p["datum"], p["typ"], p["text"] or None,
                 (f"Excel-Import: {datei.filename}, Blatt {p['blatt']}, "
                  f"Zeile {p['zeile']}"
                  + (f", Beleg {p['beleg_nr']}" if p["beleg_nr"] else ""))),
            )
            buchung_id = cur.lastrowid
            for name, betrag in p["zeilen"]:
                con.execute(
                    "INSERT INTO buchungszeile(buchung_id, kategorie_id, "
                    "betrag_cent) VALUES(?,?,?)",
                    (buchung_id, kat_vorhanden[name.strip().lower()], betrag))
        con.commit()
    except sqlite3.Error as e:
        con.rollback()
        raise HTTPException(400, f"Import abgebrochen, nichts gespeichert: {e}")

    bericht["eingespielt"] = len(neu)
    return bericht
