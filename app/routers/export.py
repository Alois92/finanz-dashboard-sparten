"""XLSX-Export und druckoptimierter Jahresbericht."""
import datetime as dt
import html
import io
import re
import sqlite3
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from ..db import db_dep

router = APIRouter(tags=["export"])
EURO_FORMAT = '#.##0,00 \u20ac'

def _xlsx_text(value):
    """Schreibt nutzerkontrollierte Texte als Literal statt als Excel-Formel."""
    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _validate(con, von=None, bis=None, sparte_id=None):
    dates = []
    for value, field in ((von, "von"), (bis, "bis")):
        if not value:
            dates.append(None)
            continue
        try:
            dates.append(dt.date.fromisoformat(value).isoformat())
        except ValueError:
            raise HTTPException(400, f"{field} muss ein gueltiges Datum im Format JJJJ-MM-TT sein")
    von, bis = dates
    if von and bis and von > bis:
        raise HTTPException(400, "von darf nicht nach bis liegen")
    if sparte_id is not None and not con.execute(
            "SELECT 1 FROM sparte WHERE id=?", (sparte_id,)).fetchone():
        raise HTTPException(404, "Sparte nicht gefunden")
    return von, bis

def _where(von, bis, sparte_id, alias="v"):
    clauses, params = [], []
    for value, expression in ((von, f"{alias}.datum>=?"), (bis, f"{alias}.datum<=?"),
                              (sparte_id, f"{alias}.sparte_id=?")):
        if value is not None:
            clauses.append(expression)
            params.append(value)
    return (" WHERE " + " AND ".join(clauses) if clauses else ""), params

def _sheet(sheet, headers, widths, euro_from=None):
    sheet.insert_rows(1)
    for index, value in enumerate(headers, 1):
        sheet.cell(1, index, value)
    fill = PatternFill("solid", fgColor="1E6E4E")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    if euro_from:
        for row in sheet.iter_rows(min_row=2):
            for index in range(euro_from, len(headers) + 1):
                row[index - 1].number_format = EURO_FORMAT
    sheet.auto_filter.ref = sheet.dimensions

@router.get("/api/export/xlsx")
def export_xlsx(von: str | None = None, bis: str | None = None,
                sparte_id: int | None = None, con: sqlite3.Connection = Depends(db_dep)):
    von, bis = _validate(con, von, bis, sparte_id)
    where, params = _where(von, bis, sparte_id)
    rows = con.execute(
        "SELECT v.datum,s.name sparte,k.name kategorie,v.typ,COALESCE(b.text,'') text,"
        "COALESCE(ko.name,'') kontakt,v.betrag_cent FROM v_einnahmen_ausgaben v "
        "JOIN buchung b ON b.id=v.buchung_id JOIN sparte s ON s.id=v.sparte_id "
        "JOIN kategorie k ON k.id=v.kategorie_id LEFT JOIN kontakt ko ON ko.id=b.kontakt_id" +
        where + " ORDER BY v.datum,v.buchung_id,v.zeile_id", params).fetchall()
    months = con.execute(
        "SELECT strftime('%Y-%m',v.datum) monat,"
        "COALESCE(SUM(CASE WHEN v.typ='einnahme' THEN v.betrag_cent END),0) ein,"
        "COALESCE(SUM(CASE WHEN v.typ='ausgabe' THEN v.betrag_cent END),0) aus "
        "FROM v_einnahmen_ausgaben v" + where + " GROUP BY monat ORDER BY monat",
        params).fetchall()
    cats = con.execute(
        "SELECT s.name sparte,k.name kategorie,"
        "COALESCE(SUM(CASE WHEN v.typ='einnahme' THEN v.betrag_cent END),0) ein,"
        "COALESCE(SUM(CASE WHEN v.typ='ausgabe' THEN v.betrag_cent END),0) aus "
        "FROM v_einnahmen_ausgaben v JOIN sparte s ON s.id=v.sparte_id "
        "JOIN kategorie k ON k.id=v.kategorie_id" + where +
        " GROUP BY s.id,k.id ORDER BY s.sortierung,k.sortierung,k.name", params).fetchall()
    wb = Workbook(); bookings = wb.active; bookings.title = "Buchungen"
    skipped = 0
    for row in rows:
        try:
            bookings.append([_xlsx_text(row["datum"]), _xlsx_text(row["sparte"]),
                             _xlsx_text(row["kategorie"]), _xlsx_text(row["typ"]),
                             _xlsx_text(row["text"]), _xlsx_text(row["kontakt"]), row["betrag_cent"]/100])
        except (TypeError, ValueError):
            skipped += 1
    _sheet(bookings, ["Datum","Sparte","Kategorie","Typ","Text","Kontakt","Betrag \u20ac"],
           (12,24,28,12,36,24,16), 7)
    monthly = wb.create_sheet("Monatssummen")
    for row in months:
        income, expense = row["ein"] or 0, row["aus"] or 0
        monthly.append([row["monat"],income/100,expense/100,(income-expense)/100])
    _sheet(monthly, ["Monat","Einnahmen","Ausgaben","Saldo"], (14,18,18,18), 2)
    category = wb.create_sheet("Kategorien"); totals, overall = {}, [0, 0]
    for row in cats:
        income, expense = row["ein"] or 0, row["aus"] or 0
        category.append([_xlsx_text(row["sparte"]), _xlsx_text(row["kategorie"]),
                         income/100, expense/100, (income-expense)/100])
        total = totals.setdefault(row["sparte"], [0, 0])
        total[0] += income; total[1] += expense
        overall[0] += income; overall[1] += expense
    for name, (income, expense) in totals.items():
        category.append([_xlsx_text(name),"Gesamt",income/100,expense/100,(income-expense)/100])
    category.append(["Gesamt","Gesamt",overall[0]/100,overall[1]/100,
                     (overall[0]-overall[1])/100])
    _sheet(category, ["Sparte","Kategorie","Einnahmen","Ausgaben","Saldo"],
           (24,30,18,18,18), 3)
    wb.properties.description = f"Uebersprungene Datensaetze: {skipped}"
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return StreamingResponse(stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="finanz-export.xlsx"'})

def _euro(cents):
    value = f"{abs(cents)/100:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return ("-" if cents < 0 else "") + value + " \u20ac"

def _sums(con, start, end, sid):
    where, params = _where(start, end, sid)
    row = con.execute(
        "SELECT COALESCE(SUM(CASE WHEN v.typ='einnahme' THEN v.betrag_cent END),0) ein,"
        "COALESCE(SUM(CASE WHEN v.typ='ausgabe' THEN v.betrag_cent END),0) aus "
        "FROM v_einnahmen_ausgaben v" + where, params).fetchone()
    return row["ein"] or 0, row["aus"] or 0

@router.get("/export/bericht", response_class=HTMLResponse)
def jahresbericht(jahr: str, sparte_id: int | None = None,
                  con: sqlite3.Connection = Depends(db_dep)):
    if not re.fullmatch(r"\d{4}", jahr) or not 1900 <= int(jahr) <= 9999:
        raise HTTPException(400, "jahr muss vierstellig sein")
    start, end = f"{jahr}-01-01", f"{jahr}-12-31"
    _validate(con, start, end, sparte_id)
    if sparte_id is None:
        divisions = con.execute(
            "SELECT id,name FROM sparte WHERE aktiv=1 ORDER BY sortierung,name").fetchall()
    else:
        divisions = con.execute("SELECT id,name FROM sparte WHERE id=?", (sparte_id,)).fetchall()
    total_in, total_out = _sums(con, start, end, sparte_id)
    sections = []
    for division in divisions:
        income, expense = _sums(con, start, end, division["id"])
        monthly = con.execute(
            "SELECT CAST(strftime('%m',v.datum) AS INTEGER) monat,"
            "COALESCE(SUM(CASE WHEN v.typ='einnahme' THEN v.betrag_cent END),0) ein,"
            "COALESCE(SUM(CASE WHEN v.typ='ausgabe' THEN v.betrag_cent END),0) aus "
            "FROM v_einnahmen_ausgaben v WHERE v.datum>=? AND v.datum<=? "
            "AND v.sparte_id=? GROUP BY monat ORDER BY monat",
            (start,end,division["id"])).fetchall()
        by_month = {row["monat"]:(row["ein"] or 0,row["aus"] or 0) for row in monthly}
        month_html = "".join(
            f"<tr><td>{month:02d}</td><td>{_euro(by_month.get(month,(0,0))[0])}</td>"
            f"<td>{_euro(by_month.get(month,(0,0))[1])}</td>"
            f"<td>{_euro(by_month.get(month,(0,0))[0]-by_month.get(month,(0,0))[1])}</td></tr>"
            for month in range(1,13))
        cats = con.execute(
            "SELECT k.name,COALESCE(SUM(CASE WHEN v.typ='einnahme' THEN v.betrag_cent END),0) ein,"
            "COALESCE(SUM(CASE WHEN v.typ='ausgabe' THEN v.betrag_cent END),0) aus "
            "FROM v_einnahmen_ausgaben v JOIN kategorie k ON k.id=v.kategorie_id "
            "WHERE v.datum>=? AND v.datum<=? AND v.sparte_id=? "
            "GROUP BY k.id ORDER BY k.sortierung,k.name",(start,end,division["id"])).fetchall()
        cat_html = "".join(
            f"<tr><td>{html.escape(row['name'])}</td><td>{_euro(row['ein'] or 0)}</td>"
            f"<td>{_euro(row['aus'] or 0)}</td>"
            f"<td>{_euro((row['ein'] or 0)-(row['aus'] or 0))}</td></tr>"
            for row in cats) or '<tr><td colspan="4">Keine Buchungen</td></tr>'
        sections.append(
            f'<section class="division"><h2>{html.escape(division["name"])}</h2>'
            f'<p><b>Einnahmen {_euro(income)}</b> &middot; <b>Ausgaben {_euro(expense)}</b> &middot; '
            f'<b>Saldo {_euro(income-expense)}</b></p><h3>Monatssummen</h3>'
            f'<table><tr><th>Monat</th><th>Einnahmen</th><th>Ausgaben</th><th>Saldo</th></tr>'
            f'{month_html}</table><h3>Kategorien</h3><table><tr><th>Kategorie</th>'
            f'<th>Einnahmen</th><th>Ausgaben</th><th>Saldo</th></tr>{cat_html}</table></section>')
    page = f"""<!doctype html><html lang="de"><head><meta charset="utf-8">
<title>Jahresbericht {jahr}</title><style>
body{{font:14px Arial;color:#17221c;margin:32px}}h1{{font-size:34px}}
table{{width:100%;border-collapse:collapse}}
th,td{{border-bottom:1px solid #ccd5d0;padding:6px;text-align:right}}
th:first-child,td:first-child{{text-align:left}}
.cover{{min-height:90vh;display:flex;flex-direction:column;justify-content:center}}
.division{{page-break-before:always}}.toolbar{{position:fixed;right:24px;top:18px}}
@page{{size:A4;margin:16mm}}@media print{{.toolbar{{display:none}}body{{margin:0}}}}
</style></head><body>
<div class="toolbar"><button onclick="window.print()">Drucken / Als PDF speichern</button></div>
<section class="cover"><h1>Jahresbericht {jahr}</h1><p>Finanz-Dashboard Hohenegg</p>
<p><b>Einnahmen {_euro(total_in)}</b> &middot; <b>Ausgaben {_euro(total_out)}</b> &middot;
<b>Saldo {_euro(total_in-total_out)}</b></p></section>
{''.join(sections)}</body></html>"""
    return HTMLResponse(page)
